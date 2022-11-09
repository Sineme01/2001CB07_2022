

from datetime import datetime
start_time = datetime.now()

#importing libraries and module

import csv
from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter as l_g
import pandas as pd
from openpyxl.styles.borders import Border
from openpyxl.styles import borders

# function to calclulate rank of octant in particular range
def changeArr(input1):

	# Copy input array into newArray
	newArray = input1.copy()
	
	# Sort newArray[] in ascending order
	newArray.sort()
	
	# Dictionary to store the rank of
	# the array element
	ranks = {}
	
	
	rank = 1
	for index in range(len(newArray)):
		element = newArray[index];
	
		# Update rank of element
		if element not in ranks:
			ranks[element] = rank
			rank += 1
		
	# Assign ranks to elements
	for index in range(len(input1)):
		element = input1[index]
        
		input1[index] = 9-ranks[input1[index]]
      
            
            
            


#Help https://youtu.be/N6PBd4XdnEw
def octant_range_names(mod=5000):
    octant_name_id_mapping = {1:"Internal outward interaction", -1:"External outward interaction", 2:"External Ejection", -2:"Internal Ejection", 3:"External inward interaction", -3:"Internal inward interaction", 4:"Internal sweep", -4:"External sweep"}
    octant_rank_id_mapping = {0:1, 1:-1, 2:2, 3:-2, 4:3, 5:-3, 6:4, 7:-4}
    count_rank_id_mapping = {0:0, 1:0, 2:0, 3:0, 4:0, 5:0, 6:0, 7:0}
    #use pandas library to read data
    data_frame=pd.read_excel("octant_input.xlsx")
    
    #use load_workbook library to work with excel
    work_book=load_workbook("octant_input.xlsx")
    
    #work with active or current sheet
    work_sheet=work_book.active
    len_frame=len(data_frame['Time'].tolist())

    #store u,v,w in list
    list_u1=data_frame['U'].tolist()
    list_v1=data_frame['V'].tolist()
    list_w1=data_frame['W'].tolist()
    c=[list_u1,list_v1,list_w1]

    #calculate mean of u,v,w
    mean_u=data_frame['U'].mean()
    mean_v=data_frame['V'].mean()
    mean_w=data_frame['W'].mean()

    row0=["U Avg","V Avg","W Avg"]
    for i in range(3):
        work_sheet[l_g(i+5)+"1"]=row0[i]

    mean=[mean_u,mean_v,mean_w]
    for i in range(3):
        work_sheet[l_g(i+5)+"2"]=mean[i]

    b_list=["U'=U-Uavg","V'=V-Vavg","W'=W-Wavg"]
    for i in range(3):
        work_sheet[l_g(i+8)+"1"]=b_list[i]

    for i in range(3):
        for j in range(1,len_frame+1):
            work_sheet[l_g(i+8)+str(j+1)]=c[i][j-1]-mean[i]

    dint = {'+++': "+1", "++-": "-1", "-++": "+2", "-+-": "-2",
            "--+": "+3", "---": "-3", "+-+": "+4", "+--": "-4"}

    work_sheet[l_g(11)+"1"]="Octant"

    # calculating oct_ant value

    for i in range(1,len_frame+1):
        xim=""
        for j in range(8,11):
            if(work_sheet[l_g(j)+str(i+1)].value>=0):
                xim+='+'
            else:
                xim+='-'
        work_sheet[l_g(11)+str(i+1)]=dint[xim]
    
    fint1=[1, -1, 2 ,-2, 3, -3, 4, 4]
    fint=["+1","-1","+2","-2","+3","-3","+4","-4","Rank 1","Rank 2","Rank 3","Rank 4","Rank 5","Rank 6","Rank 7","Rank 8","Rank1 Octant ID","Rank1 Octant Name"]
    work_sheet[l_g(13)+"1"]="octant ID"
    for i in range(len(fint)):
        work_sheet[l_g(i+14)+"1"]=fint[i]
    work_sheet[l_g(13)+"2"]="Overall Count"
    count=0
    for i in range(1,len_frame+1):
        count= count+1    
    oct_ant=[]

    for i in range(1,len_frame+1):
        oct_ant.append(work_sheet[l_g(11)+str(i+1)].value)
    lll=[]
    for i in range(8):
        work_sheet[l_g(14+i)+"2"]=oct_ant.count(fint[i])
        lll.append(oct_ant.count(fint[i]))
    changeArr(lll)
    temp=0
    for j in range(8):    
        work_sheet[l_g(22+j)+"2"]=lll[j]
        if lll[j]==1 : temp=j
    work_sheet[l_g(30)+'2']=octant_rank_id_mapping[temp]
    work_sheet[l_g(31)+'2']=octant_name_id_mapping[octant_rank_id_mapping[temp]]
    work_sheet[l_g(12)+'3']="User Input"

    total=mod
    work_sheet[l_g(13)+'3']=f"Mod {total}"
    row_curr=4

    #calculating oct_ant count for different intervals and rank of their octant
    for i in range(0,30000,total):
        if(i+total>=len_frame):
            xim=len_frame
        else:
            xim=i+total-1
        work_sheet[l_g(13)+str(row_curr)]=f"{i}-{xim}"
        v=oct_ant[i:i+total]
        ccc=[]
        for j in range(8):
            ccc.append(v.count(fint[j]))
            work_sheet[l_g(14+j)+str(row_curr)]=v.count(fint[j])
        changeArr(ccc)
        temp =0
        for j in range(8):
            if ccc[j]==1 : 
                temp=j
                count_rank_id_mapping[temp]+=1
            work_sheet[l_g(22+j)+str(row_curr)]=ccc[j]
        work_sheet[l_g(30)+str(row_curr)]=octant_rank_id_mapping[temp]
        work_sheet[l_g(31)+str(row_curr)]=octant_name_id_mapping[octant_rank_id_mapping[temp]]
        row_curr+=1


    row_curr+=3

    
    # code for count of rank 1 mod values
    gint=['Octant ID',"Octant Name","Count of Rank1 Mod value"]
    for i in range(3):
        work_sheet[l_g(14+i)+str(row_curr)]=gint[i]
    row_curr+=1
    for i in range(8):
        work_sheet[l_g(14)+str(row_curr)]=octant_rank_id_mapping[i]
        work_sheet[l_g(15)+str(row_curr)]=octant_name_id_mapping[octant_rank_id_mapping[i]]
        work_sheet[l_g(16)+str(row_curr)]=count_rank_id_mapping[i]
        row_curr+=1
        #creating output file
    
    #inserting the required 1st row
    work_sheet.insert_rows(1)
    for i in range(len(fint1)):
        work_sheet[l_g(i+22)+"1"]=fint1[i]
    work_book.save("octant_output_ranking_excel.xlsx")


mod=5000 
octant_range_names(mod)



#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
