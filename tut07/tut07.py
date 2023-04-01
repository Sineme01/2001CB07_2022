# importing modules.
from platform import python_version
from openpyxl import workbook, load_workbook
from openpyxl.styles.borders import Border
from openpyxl.styles import borders
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter as gl
from datetime import datetime
import pandas as pd
import os
start_time = datetime.now()


def octant_analysis(mod):
    # code to each file in directory
    out_path = 'output'
    curr_dic = os.getcwd()
    path = f'{curr_dic}/input'
    os.chdir(path)
    file_list = os.listdir()
    os.chdir(curr_dic)
    os.mkdir(out_path)
    for file in file_list:
        if file.endswith(".xlsx"):
            x = file.split('.')
            File__name = ''
            if (x[1] != 'xlsx'):
                File__name = x[0]+'.'+x[1]
            else:
                File__name = x[0]

            File_location = f"{path}/{file}"
            bita = 0
            wb1 = load_workbook(File_location)

            wsheet = wb1.active

            fill_cell = PatternFill(patternType='solid',
                                    fgColor='FFFF00')
            add_col = 33

            def func_first():

                Octant__Name_Id_Map = {"+1": "Internal outward interaction", "-1": "External outward interaction", "+2": "External Ejection",
                                       "-2": "Internal Ejection", "+3": "External inward interaction", "-3": "Internal inward interaction", "+4": "Internal sweep", "-4": "External sweep"}

                # pandas library is used to read data.
                dataframe = pd.read_excel(File_location)
                # load_workbook library is used to work with excel.
                length = len(dataframe['T'].tolist())
                # store u,v,w in list
                u1 = dataframe['U'].tolist()
                v1 = dataframe['V'].tolist()
                w1 = dataframe['W'].tolist()
                a = ["U Avg", "V Avg", "W Avg"]  # list
                # calculate the mean of u,v,w
                u = dataframe['U'].mean()
                v = dataframe['V'].mean()
                w = dataframe['W'].mean()
                mean = [u, v, w]  # list for storing means
                for i in range(3):
                    wsheet[gl(i+5)+"1"] = a[i]
                for i in range(3):
                    wsheet[gl(i+5)+"2"] = '%.3f' % mean[i]
                b = ["U'=U-Uavg", "V'=V-Vavg", "W'=W-Wavg"]
                c = [u1, v1, w1]
                for i in range(3):
                    wsheet[gl(i+8)+"1"] = b[i]
                for i in range(3):
                    for j in range(1, length+1):
                        wsheet[gl(i+8)+str(j+1)
                               ] = round((c[i][j-1]-mean[i]), 3)
                d = {'+++': "+1", "++-": "-1", "-++": "+2", "-+-": "-2",
                     "--+": "+3", "---": "-3", "+-+": "+4", "+--": "-4"}
                wsheet[gl(11)+"1"] = "Octant"
                # loop to calculate octant value
                for i in range(1, length+1):
                    x = ""
                    for j in range(8, 11):
                        if (wsheet[gl(j)+str(i+1)].value >= 0):
                            x += '+'
                        else:
                            x += '-'
                    wsheet[gl(11)+str(i+1)] = d[x]
                wsheet[gl(13)+"1"] = "Overall Octant Count"
                f = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
                for i in range(len(f)):
                    wsheet[gl(i+14)+"3"] = f[i]
                wsheet[gl(13)+"3"] = "Octant ID"
                wsheet[gl(12)+"4"] = f"Mod{mod}"

                wsheet[gl(13)+"4"] = "Overall Count"
                octant = []
                for i in range(1, length+1):
                    octant.append(wsheet[gl(11)+str(i+3)].value)

                for i in range(8):
                    wsheet[gl(14+i)+"4"] = octant.count(f[i])
                t = mod
                current_row = 5
                # calculate octant count in different intervals
                cou = 0
                for i in range(0, len(u1), t):
                    cou += 1
                    if (i+t >= length):
                        x = length
                    else:
                        x = i+t-1
                    wsheet[gl(13)+str(current_row)] = f"{i}-{x}"
                    v = octant[i:i+t]
                    for j in range(8):
                        wsheet[gl(14+j)+str(current_row)] = v.count(f[j])
                    current_row += 1
                current_row = 1
                current_row += 2
                for i in range(8):
                    wsheet[gl(22+i)+str(current_row)] = f"Rank of {f[i]}"
                current_row += 1
                x = []
                for i in range(8):
                    x.append(
                        [wsheet[gl(14+i)+str(current_row)].value, f[i]])
                x.sort(reverse=True)
                d = {"+1": 22, "-1": 23, "+2": 24, "-2": 25,
                     "+3": 26, "-3": 27, "+4": 28, "-4": 29}
                # insert rank values
                for i in range(8):
                    if (i+1 == 1):

                        wsheet[gl(d[x[i][1]])+str(current_row)] = i+1
                        wsheet[gl(d[x[i][1]])+str(current_row)
                               ].fill = fill_cell
                    else:
                        wsheet[gl(d[x[i][1]])+str(current_row)] = i+1

                current_row += 1
                g = cou
                while (cou > 0):
                    x = []
                    for i in range(8):
                        x.append(
                            [wsheet[gl(14+i)+str(current_row)].value, f[i]])
                    x.sort(reverse=True)
                    for i in range(8):
                        if (i+1 == 1):

                            wsheet[gl(d[x[i][1]])+str(current_row)] = i+1
                            wsheet[gl(d[x[i][1]])+str(current_row)
                                   ].fill = fill_cell
                        else:
                            wsheet[gl(d[x[i][1]])+str(current_row)] = i+1

                    current_row += 1
                    cou -= 1

                current_row = 3
                wsheet[gl(30)+str(current_row)] = "Rank1 Octant ID"
                wsheet[gl(31)+str(current_row)] = "Rank1 Octant Name"
                current_row += 1
                x = []
                for i in range(8):
                    x.append(
                        [wsheet[gl(14+i)+str(current_row)].value, f[i]])
                x.sort(reverse=True)
                wsheet[gl(30)+str(current_row)] = x[0][1]
                current_row += 1
                p = g
                y = {"+1": 0, "-1": 0, "+2": 0, "-2": 0,
                     "+3": 0, "-3": 0, "+4": 0, "-4": 0}
                while (g > 0):

                    x = []
                    for i in range(8):
                        x.append(
                            [wsheet[gl(14+i)+str(current_row)].value, f[i]])
                    x.sort(reverse=True)
                    y[x[0][1]] += 1
                    wsheet[gl(30)+str(current_row)] = x[0][1]
                    current_row += 1
                    g -= 1
                k = Octant__Name_Id_Map
                current_row = 4
                wsheet[gl(31)+str(current_row)
                       ] = k[wsheet[gl(30)+str(current_row)].value]
                current_row += 1
                while (p > 0):
                    wsheet[gl(31)+str(current_row)
                           ] = k[wsheet[gl(30)+str(current_row)].value]
                    current_row += 1
                    p -= 1
                current_row += 3
                fr = current_row
                wsheet[gl(13+15)+str(current_row)] = 'Octant ID'
                wsheet[gl(13+16)+str(current_row)] = 'Octant Name'
                wsheet[gl(13+17)+str(current_row)
                       ] = 'Count of Rank1 Mod Values'
                current_row += 1
                for i in range(8):
                    wsheet[gl(13+15)+str(current_row)] = f[i]
                    wsheet[gl(13+16)+str(current_row)] = k[f[i]]
                    wsheet[gl(13+17)+str(current_row)] = y[f[i]]
                    current_row += 1

                for idx, col in enumerate(wsheet.columns, 1):
                    wsheet.column_dimensions[gl(idx)].auto_size = True
                border1 = borders.Side(
                    style=None, color='FF000000', border_style='thin')
                border0 = borders.Side(
                    style=None, color='FF000000', border_style='thin')
                thin = Border(left=border1, right=border0,
                              bottom=border0, top=border0)

                for row in wsheet.iter_rows(min_row=3, min_col=13, max_row=8, max_col=31):
                    for cell in row:
                        cell.border = thin
                for row in wsheet.iter_rows(min_row=12, min_col=28, max_row=20, max_col=30):
                    for cell in row:
                        cell.border = thin

            def func_second():

                # use pandas library to read data
                dataframe = pd.read_excel(File_location)
                # use load_workbook library to work with excel

                # work with active or current sheet
                # wsheet = wb1.active
                length = len(dataframe['T'].tolist())
                # calculate mean of u,v,w
                u = dataframe['U'].mean()
                v = dataframe['V'].mean()
                w = dataframe['W'].mean()
                # store u,v,w in list
                u1 = dataframe['U'].tolist()
                v1 = dataframe['V'].tolist()
                w1 = dataframe['W'].tolist()
                a = ["U Avg", "V Avg", "W Avg"]
                b = ["U'=U-Uavg", "V'=V-Vavg", "W'=W-Wavg"]
                c = [u1, v1, w1]
                d = {'+++': "+1", "++-": "-1", "-++": "+2", "-+-": "-2",
                     "--+": "+3", "---": "-3", "+-+": "+4", "+--": "-4"}
                for i in range(3):
                    wsheet[gl(add_col+i+5)+"1"] = a[i]
                mean = [u, v, w]
                for i in range(3):
                    wsheet[gl(add_col+i+5)+"2"] = mean[i]
                for i in range(3):
                    wsheet[gl(add_col+i+8)+"1"] = b[i]
                for i in range(3):
                    for j in range(1, length+1):
                        wsheet[gl(add_col+i+8)+str(j+1)
                               ] = c[i][j-1]-mean[i]
                wsheet[gl(add_col+11)+"1"] = "Octat"
                # loop to calculate octant value
                for i in range(1, length+1):
                    x = ""
                    for j in range(8, 11):
                        if (wsheet[gl(add_col+j)+str(i+1)].value >= 0):
                            x += '+'
                        else:
                            x += '-'
                    wsheet[gl(add_col+11)+str(i+1)] = d[x]

                f = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]

                octant = []
                for i in range(1, length+1):
                    octant.append(wsheet[gl(add_col+11)+str(i+1)].value)

                t = mod

                d1 = {}
                g = ['Count', "+1", "-1", "+2",
                     "-2", "+3", "-3", "+4", "-4"]
                # calculate Overall Transition
                current_row = 1
                wsheet[gl(add_col+13)+str(current_row)
                       ] = "Overall Transition Count"
                current_row += 1
                wsheet[gl(add_col+14)+str(current_row)] = "To"
                current_row += 1

                for k in range(9):
                    wsheet[gl(add_col+13+k)+str(current_row)] = g[k]
                for k in range(9):
                    wsheet[gl(add_col+13)+str(current_row+k)] = g[k]
                current_row += 1
                wsheet[gl(add_col+12)+str(current_row)] = "From"
                s = current_row
                for p in range(8):
                    for r in range(8):
                        d1[f[p]+f[r]] = gl(add_col+14+r)+str(s)

                    s += 1

                current_row += 1
                s = current_row

                v = octant

                for j in range(len(v)-1):
                    x = v[j]+v[j+1]
                    if (wsheet[d1[x]].value):

                        wsheet[d1[x]] = wsheet[d1[x]].value+1

                    else:
                        wsheet[d1[x]] = 1

                m = -1

                row = 4
                for ga in range(8):
                    for j in range(8):
                        try:
                            m = max(
                                m, wsheet[gl(add_col+14+j)+str(row)].value)
                        except:
                            continue
                    for k in range(8):
                        if (wsheet[gl(add_col+14+k)+str(row)].value == m):
                            wsheet[gl(add_col+14+k)+str(row)
                                   ].fill = fill_cell
                    m = -1
                    row += 1

                border1 = borders.Side(
                    style=None, color='FF000000', border_style='thin')
                border0 = borders.Side(
                    style=None, color='FF000000', border_style='thin')
                thin = Border(left=border1, right=border0,
                              bottom=border0, top=border0)

                for row in wsheet.iter_rows(min_row=current_row-2, min_col=add_col+13, max_row=current_row+6, max_col=add_col+13+8):
                    for cell in row:
                        cell.border = thin
                current_row += 9

                # Calculate Transition in different intervals
                for i in range(0, length, t):
                    d1 = {}

                    wsheet[gl(add_col+13)+str(current_row)
                           ] = "Mod Transition Count"
                    current_row += 1
                    if (i+t >= length):
                        x = length
                    else:
                        x = i+t-1
                    wsheet[gl(add_col+13)+str(current_row)] = f"{i}-{x}"
                    wsheet[gl(add_col+14)+str(current_row)] = "To"
                    current_row += 1
                    for k in range(9):
                        wsheet[gl(add_col+13+k)+str(current_row)] = g[k]
                    for k in range(9):
                        wsheet[gl(add_col+13)+str(current_row+k)] = g[k]
                    current_row += 1
                    wsheet[gl(add_col+12)+str(current_row)] = "From"
                    ro = current_row
                    s = current_row
                    for p in range(8):
                        for r in range(8):
                            d1[f[p]+f[r]] = gl(add_col+14+r)+str(s)
                        s += 1
                    v = octant[i:i+t]
                    border1 = borders.Side(
                        style=None, color='FF000000', border_style='thin')
                    border0 = borders.Side(
                        style=None, color='FF000000', border_style='thin')
                    thin = Border(left=border1, right=border0,
                                  bottom=border0, top=border0)

                    for row in wsheet.iter_rows(min_row=current_row-1, min_col=add_col+13, max_row=current_row+7, max_col=add_col+13+8):
                        for cell in row:
                            cell.border = thin
                    current_row += 10
                    if (i+t >= length):

                        v = octant[i:i+t]
                    else:

                        v = octant[i:i+t+1]
                    for j in range(len(v)-1):
                        x = v[j]+v[j+1]
                        if (wsheet[d1[x]].value):

                            wsheet[d1[x]] = wsheet[d1[x]].value+1

                        else:
                            wsheet[d1[x]] = 1
                    m = -1

                    for ga in range(8):
                        for j in range(8):
                            try:
                                m = max(
                                    m, wsheet[gl(add_col+14+j)+str(ro)].value)
                            except:
                                continue
                        for k in range(8):
                            if (wsheet[gl(add_col+14+k)+str(ro)].value == m):
                                wsheet[gl(add_col+14+k)+str(ro)
                                       ].fill = fill_cell
                        m = -1
                        ro += 1
                wsheet.delete_cols(33, 12)

            def func_third():
                add_col = 42

                # use pandas library to read data
                dataframe = pd.read_excel(File_location)
                # use load_workbook library to work with excel

                # work with active or current sheet

                length = len(dataframe['T'].tolist())
                # calculate mean of u,v,w
                u = dataframe['U'].mean()
                v = dataframe['V'].mean()
                w = dataframe['W'].mean()
                # store u,v,w in list
                u1 = dataframe['U'].tolist()
                v1 = dataframe['V'].tolist()
                w1 = dataframe['W'].tolist()
                a = ["U Avg", "V Avg", "W Avg"]
                for i in range(3):
                    wsheet[gl(add_col+i+5)+"1"] = a[i]
                mean = [u, v, w]
                for i in range(3):
                    wsheet[gl(add_col+i+5)+"2"] = mean[i]
                b = ["U'=U-Uavg", "V'=V-Vavg", "W'=W-Wavg"]
                c = [u1, v1, w1]
                for i in range(3):
                    wsheet[gl(add_col+i+8)+"1"] = b[i]
                # calculate u',v',w'
                for i in range(3):
                    for j in range(1, length+1):
                        wsheet[gl(add_col+i+8)+str(j+1)
                               ] = c[i][j-1]-mean[i]
                d = {'+++': "+1", "++-": "-1", "-++": "+2", "-+-": "-2",
                     "--+": "+3", "---": "-3", "+-+": "+4", "+--": "-4"}
                wsheet[gl(add_col+11)+"1"] = "Octant"
                # loop to calculate octant value
                for i in range(1, length+1):
                    x = ""
                    for j in range(8, 11):
                        if (wsheet[gl(add_col+j)+str(i+1)].value >= 0):
                            x += '+'
                        else:
                            x += '-'
                    wsheet[gl(add_col+11)+str(i+1)] = d[x]
                wsheet[gl(add_col+13)+"1"] = "Longest Subsequence Length"
                wsheet[gl(add_col+13)+"3"] = "Count"
                wsheet[gl(add_col+14)+"3"] = "Longest Subsequence Length"
                wsheet[gl(add_col+15)+"3"] = "Count"
                f = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
                for i in range(len(f)):
                    wsheet[gl(add_col+13)+str(i+4)] = f[i]
                octant = []
                for i in range(1, length+1):
                    octant.append(wsheet[gl(add_col+11)+str(i+1)].value)
                current_row = 4
                time = dataframe['T'].tolist()
                # list for time ranges
                rangel = [[] for i in range(8)]
                # loop to calculate maximum subsequence length and its count
                oc = 0
                for i in f:
                    x = 0
                    cl = 0
                    cou = 0
                    ra = []
                    for j in range(len(octant)):
                        if (i == octant[j]):
                            ra.append(time[j])
                            cl += 1
                        else:

                            if (x == cl and cl > 0):
                                rangel[oc].append([ra[0], ra[-1]])
                                cou += 1

                            if (cl > x):
                                if (len(rangel[oc]) != 0):
                                    while (len(rangel[oc]) > 0):
                                        rangel[oc].pop()

                                rangel[oc].append([ra[0], ra[-1]])
                                cou = 1
                            x = max(cl, x)
                            cl = 0
                            ra = []

                    if (x == cl and cl > 0):
                        rangel[oc].append([ra[0], ra[-1]])
                        cou += 1

                    if (cl > x):
                        if (len(rangel[oc]) != 0):
                            while (len(rangel[oc]) > 0):
                                rangel[oc].pop()

                        rangel[oc].append([ra[0], ra[-1]])
                        cou = 1
                        x = max(cl, x)
                        cl = 0
                    oc += 1

                    wsheet[gl(add_col+14)+str(current_row)] = x
                    wsheet[gl(add_col+15)+str(current_row)] = cou
                    current_row += 1
                # form table with border design
                border1 = borders.Side(
                    style=None, color='FF000000', border_style='thin')
                border0 = borders.Side(
                    style=None, color='FF000000', border_style='thin')
                thin = Border(left=border1, right=border0,
                              bottom=border0, top=border0)

                for row in wsheet.iter_rows(min_row=3, min_col=add_col+13, max_row=current_row-1, max_col=add_col+15):
                    for cell in row:
                        cell.border = thin

                # make another table with time range values
                wsheet[gl(add_col+17) +
                       "1"] = "Longest Subsequence Length with Range"
                wsheet[gl(add_col+17)+"3"] = "Count"
                wsheet[gl(add_col+18)+"3"] = "Longest Subsequence Length"
                wsheet[gl(add_col+19)+"3"] = "Count"
                current_row = 4
                insert = []
                for i in range(8):
                    insert.append([wsheet[gl(add_col+13)+str(current_row)].value, wsheet[gl(add_col+14)+str(
                        current_row)].value, wsheet[gl(add_col+15)+str(current_row)].value])
                    current_row += 1

                current_row = 3
                for i in range(8):
                    current_row += 1
                    wsheet[gl(add_col+17)+str(current_row)] = insert[i][0]
                    wsheet[gl(add_col+18)+str(current_row)] = insert[i][1]
                    wsheet[gl(add_col+19)+str(current_row)] = insert[i][2]
                    current_row += 1
                    wsheet[gl(add_col+17)+str(current_row)] = "Time"
                    wsheet[gl(add_col+18)+str(current_row)] = "From"
                    wsheet[gl(add_col+19)+str(current_row)] = "To"

                    for j in range(insert[i][2]):
                        current_row += 1
                        wsheet[gl(add_col+18)+str(current_row)
                               ] = round(rangel[i][j][0], 3)
                        wsheet[gl(add_col+19)+str(current_row)
                               ] = round(rangel[i][j][1], 3)

                border1 = borders.Side(
                    style=None, color='FF000000', border_style='thin')
                border0 = borders.Side(
                    style=None, color='FF000000', border_style='thin')
                thin = Border(left=border1, right=border0,
                              bottom=border0, top=border0)

                for row in wsheet.iter_rows(min_row=3, min_col=add_col+17, max_row=current_row, max_col=add_col+19):
                    for cell in row:
                        cell.border = thin

                # resized the columns according to their content
                for idx, col in enumerate(wsheet.columns, 1):
                    wsheet.column_dimensions[gl(
                        add_col+idx)].auto_size = True

                wsheet.delete_cols(43, 11)

            func_first()
            func_second()
            func_third()
            for i in range(1, 50):
                wsheet.column_dimensions[gl(i)].width = 24
            wb1.save(
                f'{out_path}/{File__name}_octant_analysis_mod_{mod}.xlsx')
    # except:
    #     pass


ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod = 5000  # default value
octant_analysis(mod)

end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
