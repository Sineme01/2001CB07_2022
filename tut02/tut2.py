from openpyxl.styles.borders import Border
import csv
from openpyxl.styles import borders
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter as gl
import pandas as pd

# Defining the function which takes mod as its default argument.


def octant_transition_count(mod=5000):
    # Here I am using Pandas library to read the excel file "input_octant_transition_identify.xlsx".
    dataframe = pd.read_excel("input_octant_transition_identify.xlsx")
    # In order to work with excel we have to use load_workbook library
    wbook = load_workbook("input_octant_transition_identify.xlsx")
    # Now, we work with active sheet.
    wsheet = wbook.active
    size = len(dataframe['Time'].tolist())
    # list for containing uavg, vavg and wavg values.
    alist = ["U Avg", "V Avg", "W Avg"]
    u_meanlist = dataframe['U'].mean()  # calculating mean of u column.
    v_meanlist = dataframe['V'].mean()  # calculating mean of v column.
    w_menlist = dataframe['W'].mean()  # calculating mean of w column.
    u_1 = dataframe['U'].tolist()  # storing u in ---list
    v_1 = dataframe['V'].tolist()  # storing v in ---list
    w_1 = dataframe['W'].tolist()  # storing w in ---list
    for i in range(3):
        wsheet[gl(i+5)+"1"] = alist[i]
    meanval = [u_meanlist, v_meanlist, w_menlist]
    for i in range(3):
        wsheet[gl(i+5)+"2"] = meanval[i]
    blist = ["U'=U-Uavg", "V'=V-Vavg", "W'=W-Wavg"]
    clist = [u_1, v_1, w_1]
    for i in range(3):
        wsheet[gl(i+8)+"1"] = blist[i]
    for i in range(3):
        for j in range(1, size+1):
            wsheet[gl(i+8)+str(j+1)] = clist[i][j-1]-meanval[i]
    dictionary = {'+++': "+1", "++-": "-1", "-++": "+2", "-+-": "-2",
                  "--+": "+3", "---": "-3", "+-+": "+4", "+--": "-4"}
    wsheet[gl(11)+"1"] = "Octant"
    # The loop below is used to compute octant values.
    for i in range(1, size+1):
        x = ""
        for j in range(8, 11):
            if (wsheet[gl(j)+str(i+1)].value >= 0):
                x += '+'
            else:
                x += '-'
        wsheet[gl(11)+str(i+1)] = dictionary[x]
    listf = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
    for i in range(len(listf)):
        wsheet[gl(i+14)+"1"] = listf[i]
    wsheet[gl(13)+"2"] = "Overall Count"
    wbook.save("output_octant_transition_identify.xlsx")
    octant = []
    for i in range(1, size+1):
        octant.append(wsheet[gl(11)+str(i+1)].value)
    for i in range(8):
        wsheet[gl(14+i)+"2"] = octant.count(listf[i])
    wsheet[gl(12)+'3'] = "User Input"
    t = mod
    wsheet[gl(13)+'3'] = f"Mod {t}"
    present_row = 4
    # counting octant in different intervals.
    for i in range(0, 30000, t):
        if (i+t >= size):
            x = size
        else:
            x = i+t-1
        wsheet[gl(13)+str(present_row)] = f"{i}-{x}"
        v = octant[i:i+t]
        for j in range(8):
            wsheet[gl(14+j)+str(present_row)] = v.count(listf[j])
        present_row += 1
    wsheet[gl(13)+str(present_row)] = "Verified"
    # counting total sum of individual +1,-1,+2... column for the purpose of verification.
    for i in range(8):
        ch = gl(i+14)
        wsheet[gl(14+i)+str(present_row)
               ] = f'=SUM({gl(14+i)+str(4)}:{gl(14+i)+str(present_row-1)})'
    # Here is the code for designing borders.
    border1 = borders.Side(style=None, color='FF000000', border_style='thin')
    border0 = borders.Side(style=None, color='FF000000', border_style='thin')
    thin = Border(left=border1, right=border0, bottom=border0, top=border0)
    for row in wsheet.iter_rows(min_row=1, min_col=13, max_row=present_row, max_col=13+8):
        for cell in row:
            cell.border = thin
    dictionary1 = {}
    g = ['Count', "+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
    # counting the size of g.
    cnt = 0
    for it in g:
        cnt += 1
    # Here I am calculating overall transition.
    present_row += 3
    wsheet[gl(13)+str(present_row)] = "Overall Transition Count"
    present_row += 1
    wsheet[gl(14)+str(present_row)] = "To"
    present_row += 1
    wsheet[gl(12)+str(present_row)] = "From"
    s = present_row
    for p in range(8):
        for r in range(8):
            dictionary1[listf[p]+listf[r]] = gl(14+r)+str(s)
        s += 1
    present_row += 1
    s = present_row
    v = octant
    for j in range(len(v)-1):
        x = v[j]+v[j+1]
        if (wsheet[dictionary1[x]].value):
            wsheet[dictionary1[x]] = wsheet[dictionary1[x]].value+1
        else:
            wsheet[dictionary1[x]] = 1
    border1 = borders.Side(style=None, color='FF000000', border_style='thin')
    border0 = borders.Side(style=None, color='FF000000', border_style='thin')
    thin = Border(left=border1, right=border0, bottom=border0, top=border0)
    for row in wsheet.iter_rows(min_row=present_row-2, min_col=13, max_row=present_row+6, max_col=13+8):
        for cell in row:
            cell.border = thin
        for it in row:
            cnt += 1
    present_row += 9
    for i in range(0, 30000, t):
        dictionary1 = {}
        wsheet[gl(13)+str(present_row)] = "Mod Transition Count"
        present_row += 1
        if (i+t >= size):
            x = size
        else:
            x = i+t-1
        wsheet[gl(13)+str(present_row)] = f"{i}-{x}"
        wsheet[gl(14)+str(present_row)] = "To"
        present_row += 1
        for k in range(9):
            wsheet[gl(13+k)+str(present_row)] = g[k]
        for k in range(9):
            wsheet[gl(13)+str(present_row+k)] = g[k]
        present_row += 1
        wsheet[gl(12)+str(present_row)] = "From"
        s = present_row
        for p in range(8):
            for r in range(8):
                dictionary1[listf[p]+listf[r]] = gl(14+r)+str(s)
            s += 1
        voct = octant[i:i+t]
        boundary1 = borders.Side(
            style=None, color='FF000000', border_style='thin')
        boundary0 = borders.Side(
            style=None, color='FF000000', border_style='thin')
        narrow = Border(left=boundary1, right=boundary0,
                        bottom=boundary0, top=boundary0)
        for row in wsheet.iter_rows(min_row=present_row-1, min_col=13, max_row=present_row+7, max_col=13+8):
            for cell in row:
                cell.border = narrow
        present_row += 10
        for wun in range(9):
            cnt += 1
        for j in range(len(voct)-1):
            x = voct[j]+voct[j+1]
            if (wsheet[dictionary1[x]].value):
                wsheet[dictionary1[x]] = wsheet[dictionary1[x]].value+1
            else:
                wsheet[dictionary1[x]] = 1
    wbook.save("output_octant_transition_identify.xlsx")


octant_transition_count()
