from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter as gl
import pandas as pd
from openpyxl.styles.borders import Border
from openpyxl.styles import borders


def octant_transition_count(mod=5000):
    # Here I am using Pandas library to read the excel file "input_octant_transition_identify.xlsx".
    dataframe = pd.read_excel("input_octant_transition_identify.xlsx")
    # In order to work with excel we have to use load_workbook library
    wbook = load_workbook("input_octant_transition_identify.xlsx")
    # Now, we work with active sheet.
    wsheet = wbook.active
    length = len(dataframe['Time'].tolist())
    uu = dataframe['U'].mean()  # calculating mean of u column.
    vv = dataframe['V'].mean()  # calculating mean of v column.
    ww = dataframe['W'].mean()  # calculating mean of w column.
    u1 = dataframe['U'].tolist()  # storing u in ---list
    v1 = dataframe['V'].tolist()  # storing v in ---list
    w1 = dataframe['W'].tolist()  # storing w in ---list
    aa = ["U Avg", "V Avg", "W Avg"]
    for i in range(3):
        wsheet[gl(i+5)+"1"] = aa[i]
    meanval = [uu, vv, ww]
    for i in range(3):
        wsheet[gl(i+5)+"2"] = meanval[i]
    b = ["U'=U-Uavg", "V'=V-Vavg", "W'=W-Wavg"]
    c = [u1, v1, w1]
    for i in range(3):
        wsheet[gl(i+8)+"1"] = b[i]
    for i in range(3):
        for j in range(1, length+1):
            wsheet[gl(i+8)+str(j+1)] = c[i][j-1]-meanval[i]
    d = {'+++': "+1", "++-": "-1", "-++": "+2", "-+-": "-2",
         "--+": "+3", "---": "-3", "+-+": "+4", "+--": "-4"}
    wsheet[gl(11)+"1"] = "Octat"
    # The loop below is used to compute octant values.
    for i in range(1, length+1):
        x = ""
        for j in range(8, 11):
            if (wsheet[gl(j)+str(i+1)].value >= 0):
                x += '+'
            else:
                x += '-'
        wsheet[gl(11)+str(i+1)] = d[x]
    f = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
    for i in range(len(f)):
        wsheet[gl(i+14)+"1"] = f[i]
    wsheet[gl(13)+"2"] = "Overall Count"
    octant = []
    for i in range(1, length+1):
        octant.append(wsheet[gl(11)+str(i+1)].value)
    for i in range(8):
        wsheet[gl(14+i)+"2"] = octant.count(f[i])
    wsheet[gl(12)+'3'] = "User Input"
    t = mod
    wsheet[gl(13)+'3'] = f"Mod {t}"
    current_row = 4
    # counting octant in different intervals.
    for i in range(0, 30000, t):
        if (i+t >= length):
            x = length
        else:
            x = i+t-1
        wsheet[gl(13)+str(current_row)] = f"{i}-{x}"
        v = octant[i:i+t]
        for j in range(8):
            wsheet[gl(14+j)+str(current_row)] = v.count(f[j])
        current_row += 1
    wsheet[gl(13)+str(current_row)] = "Verified"
    # counting total sum of individual +1,-1,+2... column for the purpose of verification.
    for i in range(8):
        ch = gl(i+14)
        wsheet[gl(14+i)+str(current_row)
               ] = f'=SUM({gl(14+i)+str(4)}:{gl(14+i)+str(current_row-1)})'
    # Here is the code for designing borders.
    border1 = borders.Side(style=None, color='FF000000', border_style='thin')
    border0 = borders.Side(style=None, color='FF000000', border_style='thin')
    thin = Border(left=border1, right=border0, bottom=border0, top=border0)
    for row in wsheet.iter_rows(min_row=1, min_col=13, max_row=current_row, max_col=13+8):
        for cell in row:
            cell.border = thin
    d1 = {}
    g = ['Count', "+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
    #counting the length of g.
    cnt = 0
    for it in g:
        cnt += 1
    # Here I am calculating overall transition.
    current_row += 3
    wsheet[gl(13)+str(current_row)] = "Overall Transition Count"
    current_row += 1
    wsheet[gl(14)+str(current_row)] = "To"
    current_row += 1
    for k in range(9):
        wsheet[gl(13+k)+str(current_row)] = g[k]
    for k in range(9):
        wsheet[gl(13)+str(current_row+k)] = g[k]
    current_row += 1
    wsheet[gl(12)+str(current_row)] = "From"
    s = current_row
    for p in range(8):
        for r in range(8):
            d1[f[p]+f[r]] = gl(14+r)+str(s)
        s += 1
    current_row += 1
    s = current_row
    v = octant
    for j in range(len(v)-1):
        x = v[j]+v[j+1]
        if (wsheet[d1[x]].value):
            wsheet[d1[x]] = wsheet[d1[x]].value+1
        else:
            wsheet[d1[x]] = 1
    border1 = borders.Side(style=None, color='FF000000', border_style='thin')
    border0 = borders.Side(style=None, color='FF000000', border_style='thin')
    thin = Border(left=border1, right=border0, bottom=border0, top=border0)
    for row in wsheet.iter_rows(min_row=current_row-2, min_col=13, max_row=current_row+6, max_col=13+8):
        for cell in row:
            cell.border = thin
        for it in row:
            cnt += 1
    current_row += 9
    for i in range(0, 30000, t):
        d1 = {}
        wsheet[gl(13)+str(current_row)] = "Mod Transition Count"
        current_row += 1
        if (i+t >= length):
            x = length
        else:
            x = i+t-1
        wsheet[gl(13)+str(current_row)] = f"{i}-{x}"
        wsheet[gl(14)+str(current_row)] = "To"
        current_row += 1
        for k in range(9):
            wsheet[gl(13+k)+str(current_row)] = g[k]
        for k in range(9):
            wsheet[gl(13)+str(current_row+k)] = g[k]
        current_row += 1
        wsheet[gl(12)+str(current_row)] = "From"
        s = current_row
        for p in range(8):
            for r in range(8):
                d1[f[p]+f[r]] = gl(14+r)+str(s)
            s += 1
        v = octant[i:i+t]
        border1 = borders.Side(
            style=None, color='FF000000', border_style='thin')
        border0 = borders.Side(
            style=None, color='FF000000', border_style='thin')
        thin = Border(left=border1, right=border0, bottom=border0, top=border0)
        for row in wsheet.iter_rows(min_row=current_row-1, min_col=13, max_row=current_row+7, max_col=13+8):
            for cell in row:
                cell.border = thin
        current_row += 10
        for j in range(len(v)-1):
            x = v[j]+v[j+1]
            if (wsheet[d1[x]].value):
                wsheet[d1[x]] = wsheet[d1[x]].value+1
            else:
                wsheet[d1[x]] = 1
    wbook.save("output_octant_transition_identify.xlsx")


octant_transition_count()
