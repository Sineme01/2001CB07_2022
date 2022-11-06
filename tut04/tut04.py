import csv
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter as l_g
import pandas as pd
from openpyxl.styles.borders import Border
from openpyxl.styles import borders
from datetime import datetime
start_time = datetime.now()

# Help https://youtu.be/H37f_x4wAC0


def octant_longest_subsequence_count_with_range():
    # use pandas library to read data
    data_frame = pd.read_excel(
        "input_octant_longest_subsequence_with_range.xlsx")

    # use load_workbook library to work with excel
    work_book = load_workbook(
        "input_octant_longest_subsequence_with_range.xlsx")

    # work with active or current sheet
    work_sheet = work_book.active
    len_frame = len(data_frame['Time'].tolist())
    time_list = data_frame['Time'].tolist()
    # store u,v,w in list
    list_u1 = data_frame['U'].tolist()
    list_v1 = data_frame['V'].tolist()
    list_w1 = data_frame['W'].tolist()
    c = [list_u1, list_v1, list_w1]

    # calculate mean of u,v,w
    mean_u = data_frame['U'].mean()
    mean_v = data_frame['V'].mean()
    mean_w = data_frame['W'].mean()

    row0 = ["U Avg", "V Avg", "W Avg"]
    for i in range(3):
        work_sheet[l_g(i+5)+"1"] = row0[i]

    mean = [mean_u, mean_v, mean_w]
    for i in range(3):
        work_sheet[l_g(i+5)+"2"] = mean[i]

    b_list = ["U'=U-Uavg", "V'=V-Vavg", "W'=W-Wavg"]
    for i in range(3):
        work_sheet[l_g(i+8)+"1"] = b_list[i]

    for i in range(3):
        for j in range(1, len_frame+1):
            work_sheet[l_g(i+8)+str(j+1)] = c[i][j-1]-mean[i]

    dint = {'+++': "+1", "++-": "-1", "-++": "+2", "-+-": "-2",
            "--+": "+3", "---": "-3", "+-+": "+4", "+--": "-4"}

    work_sheet[l_g(11)+"1"] = "Octant"

    # calculating oct_ant value
    octant_list = []
    for i in range(1, len_frame+1):
        xim = ""
        for j in range(8, 11):
            if (work_sheet[l_g(j)+str(i+1)].value >= 0):
                xim += '+'
            else:
                xim += '-'
        work_sheet[l_g(11)+str(i+1)] = dint[xim]
        octant_list.append(dint[xim])

    gint = ['Count', "Longest Subsquence Length", "count"]
    lint = ['Time', "From", "To"]
    fint = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
    thisdict = {"+1": 0, "-1": 1, "+2": 2, "-2": 3,
                "+3": 4, "-3": 5, "+4": 6, "-4": 7, "0": 8}
    # made list of longest subsequence, countmam, and sotred end time of every longest subsequence
    longest = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    countmam = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    lount = 1
    timestore = [[], [], [], [], [], [], [], [], []]
    # algo for calculating longest count mam and storing end time
    for k in range(len(octant_list)-1):
        if octant_list[k] == octant_list[k-1]:
            lount += 1
        else:
            if lount == longest[thisdict[octant_list[k-1]]]:
                countmam[thisdict[octant_list[k-1]]] += 1
                timestore[thisdict[octant_list[k-1]]].append(time_list[k-1])
            elif lount > longest[thisdict[octant_list[k-1]]]:
                countmam[thisdict[octant_list[k-1]]] = 1
                longest[thisdict[octant_list[k-1]]] = lount
                timestore[thisdict[octant_list[k-1]]].clear()
                timestore[thisdict[octant_list[k-1]]].append(time_list[k-1])

            lount = 1

    row_curr = 2
    # creating table 1 title
    for k in range(3):
        work_sheet[l_g(13+k)+str(row_curr)] = gint[k]
    # cretaing table 2 title
    for k in range(3):
        work_sheet[l_g(17+k)+str(row_curr)] = gint[k]

    # updating table 1
    for kol in range(8):

        work_sheet[l_g(13)+str(kol+3)] = fint[kol]
        work_sheet[l_g(13+1)+str(kol+3)] = longest[kol]
        work_sheet[l_g(13+2)+str(kol+3)] = countmam[kol]

    # updating table 2 in desired way
    curr_row = 3
    for kol in range(8):

        work_sheet[l_g(17)+str(curr_row)] = fint[kol]
        work_sheet[l_g(17+1)+str(curr_row)] = longest[kol]
        work_sheet[l_g(17+2)+str(curr_row)] = countmam[kol]
        curr_row += 1
        for k in range(3):
            work_sheet[l_g(17+k)+str(curr_row)] = lint[k]
        curr_row += 1
        # pushong all time ranges
        for lol in range(len(timestore[kol])):
            work_sheet[l_g(18)+str(curr_row)] = timestore[kol][lol] - \
                ((longest[kol]-1)/100)
            work_sheet[l_g(19)+str(curr_row)] = timestore[kol][lol]
            curr_row += 1

    # creating border design
    border_0 = borders.Side(style=None, color='FF000000', border_style='thin')
    border_1 = borders.Side(style=None, color='FF000000', border_style='thin')
    thin = Border(left=border_1, right=border_0, bottom=border_0, top=border_0)

    # drawing border for table 1
    for row in work_sheet.iter_rows(min_row=row_curr, min_col=13, max_row=row_curr+8, max_col=13+2):
        for cell in row:
            cell.border = thin

    # drawing border for table 2
    for row in work_sheet.iter_rows(min_row=2, min_col=17, max_row=curr_row-1, max_col=17+2):
        for cell in row:
            cell.border = thin

    # saving workbook
    work_book.save("output_octant_longest_subsequence_with_range.xlsx")


octant_longest_subsequence_count_with_range()


# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
