from openpyxl.styles.borders import Border
import csv
from openpyxl.styles import borders
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter as gl
import pandas as pd

# Defining the function which takes mod as its default argument.


def octant_longest_subsequence_count():
    # Here I am using Pandas library to read the excel file "input_octant_transition_identify.xlsx".
    dataframe = pd.read_excel("input_octant_longest_subsequence.xlsx")
    # In order to work with excel we have to use load_workbook library
    wbook = load_workbook("input_octant_longest_subsequence.xlsx")
    # Now, we work with active sheet.
    wsheet = wbook.active
    size = len(dataframe['Time'].tolist())
    # list for containing uavg, vavg and wavg values.
    alist = ["U Avg", "V Avg", "W Avg"]
    u_meanlist = dataframe['U'].mean()  # calculating mean of u column.
    v_meanlist = dataframe['V'].mean()  # calculating mean of v column.
    w_menlist = dataframe['W'].mean()  # calculating mean of w column.
    u_1 = dataframe['U'].tolist()  # storing u in ---list
    v_1 = dataframe['V'].tolist()  # storing v in ---list
    w_1 = dataframe['W'].tolist()  # storing w in ---list
    for i in range(3):
        wsheet[gl(i+5)+"1"] = alist[i]
    meanval = [u_meanlist, v_meanlist, w_menlist]
    for i in range(3):
        wsheet[gl(i+5)+"2"] = meanval[i]
    blist = ["U'=U-Uavg", "V'=V-Vavg", "W'=W-Wavg"]
    clist = [u_1, v_1, w_1]
    for i in range(3):
        wsheet[gl(i+8)+"1"] = blist[i]
    for i in range(3):
        for j in range(1, size+1):
            wsheet[gl(i+8)+str(j+1)] = clist[i][j-1]-meanval[i]
    dictionary = {'+++': "+1", "++-": "-1", "-++": "+2", "-+-": "-2",
                  "--+": "+3", "---": "-3", "+-+": "+4", "+--": "-4"}
    wsheet[gl(11)+"1"] = "Octant"
    # The loop below is used to compute octant values.
    storeval = []  # This is a list containing the numerical values of octants.
    storeval.insert(0, 0)
    for i in range(1, size+1):
        x = ""
        for j in range(8, 11):
            if (wsheet[gl(j)+str(i+1)].value >= 0):
                x += '+'
            else:
                x += '-'
        wsheet[gl(11)+str(i+1)] = dictionary[x]
        numval = int(dictionary[x][1])
        if (dictionary[x][0] == '-'):
            numval *= -1
        storeval.append(numval)
    cnt = 0
    maxcnt = 0
    cntneg = 0
    maxcntneg = 0
    cnt2 = 0
    maxcnt2 = 0
    cntneg2 = 0
    maxcntneg2 = 0
    cnt3 = 0
    maxcnt3 = 0
    cntneg3 = 0
    maxcntneg3 = 0
    cnt4 = 0
    maxcnt4 = 0
    cntneg4 = 0
    maxcntneg4 = 0
    cntlist1 = [0]
    cntlistneg1 = [0]
    cntlist2 = [0]
    cntlistneg2 = [0]
    cntlist3 = [0]
    cntlistneg3 = [0]
    cntlist4 = [0]
    cntlistneg4 = [0]
    for it in storeval:
        if (it == 1):
            cnt += 1
        else:
            maxcnt = max(maxcnt, cnt)
            # countdictionary["+1"][maxcnt] += 1
            cntlist1.append(cnt)
            cnt = 0
        if (it == -1):
            cntneg += 1
        else:
            maxcntneg = max(maxcntneg, cntneg)
            # countdictionary["-1"][maxcntneg] += 1
            cntlistneg1.append(cntneg)
            cntneg = 0
        if (it == 2):
            cnt2 += 1
        else:
            maxcnt2 = max(maxcnt2, cnt2)
            # countdictionary["+2"][maxcnt2] += 1
            cntlist2.append(cnt2)
            cnt2 = 0
        if (it == -2):
            cntneg2 += 1
        else:
            maxcntneg2 = max(maxcntneg2, cntneg2)
            # countdictionary["-2"][maxcntneg2] += 1
            cntlistneg2.append(cntneg2)
            cntneg2 = 0
        if (it == 3):
            cnt3 += 1
        else:
            maxcnt3 = max(maxcnt3, cnt3)
            # countdictionary["+3"][maxcnt3] += 1
            cntlist3.append(cnt3)
            cnt3 = 0
        if (it == -3):
            cntneg3 += 1
        else:
            maxcntneg3 = max(maxcntneg3, cntneg3)
            # countdictionary["-3"][maxcntneg3] += 1
            cntlistneg3.append(cntneg3)
            cntneg3 = 0
        if (it == 4):
            cnt4 += 1
        else:
            maxcnt4 = max(maxcnt4, cnt4)
            # countdictionary["+4"][maxcnt4] += 1
            cntlist4.append(cnt4)
            cnt4 = 0
        if (it == -4):
            cntneg4 += 1
        else:
            maxcntneg4 = max(maxcntneg4, cntneg4)
            cntlistneg4.append(cntneg4)
            # countdictionary["-4"][maxcntneg4] += 1
            cntneg4 = 0
    wsheet[gl(13)+"1"] = "Count"
    wsheet[gl(13)+"2"] = "+1"
    wsheet[gl(13)+"3"] = "-1"
    wsheet[gl(13)+"4"] = "+2"
    wsheet[gl(13)+"5"] = "-2"
    wsheet[gl(13)+"6"] = "+3"
    wsheet[gl(13)+"7"] = "-3"
    wsheet[gl(13)+"8"] = "+4"
    wsheet[gl(13)+"9"] = "-4"
    wsheet[gl(14)+"1"] = "Longest Subsquence Length"
    wsheet[gl(14)+"2"] = str(maxcnt)
    wsheet[gl(14)+"3"] = str(maxcntneg)
    wsheet[gl(14)+"4"] = str(maxcnt2)
    wsheet[gl(14)+"5"] = str(maxcntneg2)
    wsheet[gl(14)+"6"] = str(maxcnt3)
    wsheet[gl(14)+"7"] = str(maxcntneg3)
    wsheet[gl(14)+"8"] = str(maxcnt4)
    wsheet[gl(14)+"9"] = str(maxcntneg4)
    wsheet[gl(15)+"1"] = "Count"
    val1 = 0
    for it in cntlist1:
        if (it == maxcnt):
            val1 += 1
    wsheet[gl(15)+"2"] = str(val1)
    val1 = 0
    for it in cntlistneg1:
        if (it == maxcntneg):
            val1 += 1
    wsheet[gl(15)+"3"] = str(val1)
    val1 = 0
    for it in cntlist2:
        if (it == maxcnt2):
            val1 += 1
    wsheet[gl(15)+"4"] = str(val1)
    val1 = 0
    for it in cntlistneg2:
        if (it == maxcntneg2):
            val1 += 1
    wsheet[gl(15)+"5"] = str(val1)
    val1 = 0
    for it in cntlist3:
        if (it == maxcnt3):
            val1 += 1
    wsheet[gl(15)+"6"] = str(val1)
    val1 = 0
    for it in cntlistneg3:
        if (it == maxcntneg3):
            val1 += 1
    wsheet[gl(15)+"7"] = str(val1)
    val1 = 0
    for it in cntlist4:
        if (it == maxcnt4):
            val1 += 1
    wsheet[gl(15)+"8"] = str(val1)
    val1 = 0
    for it in cntlistneg4:
        if (it == maxcntneg4):
            val1 += 1
    wsheet[gl(15)+"9"] = str(val1)
    border_0 = borders.Side(style=None, color='FF000000', border_style='thin')
    border_1 = borders.Side(style=None, color='FF000000', border_style='thin')
    thin = Border(left=border_1, right=border_0, bottom=border_0, top=border_0)
    row_curr = 1
    for row in wsheet.iter_rows(min_row=row_curr, min_col=13, max_row=row_curr+8, max_col=13+2):
        for cell in row:
            cell.border = thin
    wbook.save("output_octant_longest_subsequence.xlsx")


octant_longest_subsequence_count()
