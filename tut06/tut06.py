import xlsxwriter as xl
import pandas as pd
from openpyxl.styles.borders import Border
from openpyxl.styles import borders
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter as gl
import datetime as dt
from platform import python_version
import math as ma
import os
from datetime import datetime
start_time = datetime.now()
# importing os module
def R(): return map(int, input().split())


# This is a python program to demonstrate the use of os.mkdir() method.
path = 'output'
try:
    os.mkdir(path)  # create a folder in my directory named output.
except OSError as error:
    print(error)


def Check_Days(s):
    ps = s
    j, t = ps.split(" ")
    date, month, year = j.split("-")
    map_months = {'01': "January", '02': "February", "03": "March", "04": "April", "05": "May", "06": "June", '07': 'July', "08": "August", '09': "September", "10": 'October',
                  '11': "November", "12": "December"}

   # strptime() is another method available in DateTime which is used to format the time stamp.
    x = dt.datetime.strptime(
        f'{map_months[month]} {date}, {year}', '%B %d, %Y').strftime('%A')

    return x


def func_sort_date(date):
    dateList = list(date)
    siz = len(dateList)
    for i in range(siz):
        var = dateList[i].split("-")
        flag = var[2]+var[1]+var[0]
        dateList[i] = flag
    # sorting the date list for proper view of attendance report.
    dateList.sort()
    for i in range(siz):
        var = dateList[i][-2:]+'-'+dateList[i][-4:-2]+'-'+dateList[i][:4]
        dateList[i] = var

    return dateList


dataFrame = pd.read_csv(
    "input_registered_students.csv")
Roll = list(dataFrame['Roll No'])
name = list(dataFrame['Name'])
rool_key_map_name_store = {}
Siz = len(Roll)
for i in range(Siz):
    rool_key_map_name_store[Roll[i]] = name[i]

dict_time = {}
Attendace_Dataframe = pd.read_csv("input_attendance.csv")
t_list = list(Attendace_Dataframe['Timestamp'])
nm_list = list(Attendace_Dataframe['Attendance'])
registration__check = {}
for i in nm_list:
    dict_time[i[:8]] = []
for i in nm_list:
    registration__check[i[:8]] = 1
Siz = len(t_list)
for i in range(Siz):
    Day_list = nm_list[i][:8]
    dict_time[Day_list].append(t_list[i])
date_list = set()
for i in t_list:
    Day_list = Check_Days(i)
    if (Day_list == 'Monday' or Day_list == 'Thursday'):
        date_list.add(i.split(" ")[0])
date_list = func_sort_date(date_list)


def attendance_sheet(roll):
    attributes = ['Date', 'Roll No', 'Name', 'Total Attendance Count',
                  'Real', 'Duplicate', 'Invalid', 'Absent']
    wbook = xl.Workbook(f"{path}/{roll}.xlsx")
    ws = wbook.add_worksheet(f'{roll}')
    wbook.close()
    wbook = load_workbook(f"{path}/{roll}.xlsx")
    ws = wbook.active
    C_row = 1
    for i in range(len(attributes)):
        ws[gl(i+1)+str(C_row)] = attributes[i]
    ws[gl(2)+str(2)] = roll
    ws[gl(3)+str(2)] = rool_key_map_name_store[roll]
    C_row += 2
    for i in range(len(date_list)):
        ws[gl(1)+str(i+C_row)] = date_list[i]
    C_row = 3
    for i in range(len(date_list)):
        x = date_list[i]
        attendance_time = []
        for j in dict_time[roll]:
            y = j.split(" ")[1]
            z = j.split(" ")[0]
            if (x == z):
                attendance_time.append(y)
        inv = 0
        for k in attendance_time:
            pt = dt.datetime.strptime(k, '%H:%M')
            total_seconds = pt.second + pt.minute*60 + pt.hour*3600
            if (total_seconds < 14*3600 or total_seconds > 15*3600):
                inv += 1
        Ttl_attend = len(attendance_time)
        duplicate = 0
        real = 0
        absent = 0
        if (Ttl_attend-inv != 0):
            real = 1
            duplicate = max(duplicate, Ttl_attend-inv-1)
        else:
            absent = 1
        f = [Ttl_attend, real, duplicate, inv, absent]
        for p in range(4, 9):
            ws[gl(p)+str(C_row+i)] = f[p-4]
        wbook.save(f"{path}/{roll}.xlsx")


for i in rool_key_map_name_store:
    if (i in registration__check):
        attendance_sheet(i)

end_time = datetime.now()

print('Duration of Program Execution: {}'.format(end_time - start_time))


# Code

ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")
